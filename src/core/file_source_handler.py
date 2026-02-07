"""File source handler for CSV and Excel file uploads

Phase 13: CSV & Excel File Support

Handles file upload, validation, storage, schema inference, and DuckDB integration.
Files are stored on disk and queried via DuckDB's read_csv_auto(). Excel files are
converted to CSV using openpyxl/xlrd before loading into DuckDB.
"""
import asyncio
import hashlib
import logging
import os
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import contextlib
import csv
import tempfile

import aiofiles
import duckdb
from fastapi import UploadFile
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from src.config.settings import Settings
from src.core.file_utils import validate_file_path, sanitize_sheet_name
from src.database.models import FileSource

logger = logging.getLogger(__name__)


def excel_to_temp_csv(file_path: str, sheet_name: Optional[str] = None) -> str:
    """
    Convert an Excel file to a temporary CSV file for DuckDB ingestion.

    DuckDB 1.1.x does not include a read_excel() table function, so Excel files
    must be converted to CSV before loading.

    Args:
        file_path: Path to the Excel file (.xlsx or .xls)
        sheet_name: Sheet name to convert (defaults to first/active sheet)

    Returns:
        Path to the temporary CSV file. Caller is responsible for cleanup
        via os.unlink().
    """
    ext = Path(file_path).suffix.lower()

    if ext == '.xlsx':
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            fd, temp_path = tempfile.mkstemp(suffix='.csv')
            try:
                with os.fdopen(fd, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    for row in ws.iter_rows(values_only=True):
                        writer.writerow(row)
            except Exception:
                os.unlink(temp_path)
                raise
        finally:
            wb.close()

        return temp_path

    elif ext == '.xls':
        import xlrd

        wb = xlrd.open_workbook(file_path)
        if sheet_name:
            try:
                ws = wb.sheet_by_name(sheet_name)
            except xlrd.biffh.XLRDError:
                ws = wb.sheet_by_index(0)
        else:
            ws = wb.sheet_by_index(0)

        fd, temp_path = tempfile.mkstemp(suffix='.csv')
        try:
            with os.fdopen(fd, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                for row_idx in range(ws.nrows):
                    writer.writerow(ws.row_values(row_idx))
        except Exception:
            os.unlink(temp_path)
            raise

        return temp_path

    else:
        raise ValueError(f"Unsupported Excel format: {ext}")


class FileSourceHandler:
    """Handle file uploads and convert to queryable data sources."""

    # Mapping of extensions to file types
    EXTENSION_MAP = {
        '.csv': 'csv',
        '.xlsx': 'xlsx',
        '.xls': 'xls',
    }

    def __init__(self, settings: Optional[Settings] = None):
        """Initialize the handler with settings."""
        self.settings = settings or Settings()
        self.upload_dir = Path(self.settings.FILE_UPLOAD_DIR)
        self.max_size_bytes = self.settings.FILE_MAX_SIZE_MB * 1024 * 1024
        self.allowed_extensions = set(
            ext.strip().lower()
            for ext in self.settings.FILE_ALLOWED_TYPES.split(',')
        )
        self._ensure_upload_dir()

    def _ensure_upload_dir(self) -> None:
        """Ensure upload directory exists."""
        self.upload_dir.mkdir(parents=True, exist_ok=True)
        (self.upload_dir / 'global').mkdir(exist_ok=True)
        (self.upload_dir / 'sessions').mkdir(exist_ok=True)

    async def process_upload(
        self,
        file: UploadFile,
        db: AsyncSession,
        name: Optional[str] = None,
        session_id: Optional[str] = None,
        sheet_name: Optional[str] = None,
        is_global: bool = False,
        user_id: Optional[str] = None,
    ) -> FileSource:
        """
        Process an uploaded file and create a FileSource record.

        Args:
            file: The uploaded file
            db: Database session
            name: Display name (defaults to filename)
            session_id: Chat session ID (optional)
            sheet_name: Excel sheet name (optional)
            is_global: Whether file is available across all sessions
            user_id: User ID for ownership

        Returns:
            FileSource: The created file source record

        Raises:
            ValueError: If file validation fails
        """
        # Validate file
        is_valid, error_msg = await self.validate_file(file)
        if not is_valid:
            raise ValueError(error_msg)

        # Determine file type
        ext = Path(file.filename).suffix.lower()
        file_type = self.EXTENSION_MAP.get(ext, 'csv')

        # Create initial FileSource record with pending status
        display_name = name or Path(file.filename).stem
        table_name = self._generate_table_name(file.filename, 0)  # Temp ID

        file_source = FileSource(
            name=display_name,
            original_filename=file.filename,
            file_type=file_type,
            file_size_bytes=0,  # Will update after save
            file_path='',  # Will update after save
            duckdb_table_name=table_name,
            sheet_name=sheet_name,
            chat_session_id=session_id,
            is_global=is_global,
            user_id=user_id,
            processing_status='processing',
            expires_at=datetime.now(timezone.utc) + timedelta(days=self.settings.FILE_AUTO_CLEANUP_DAYS)
            if self.settings.FILE_AUTO_CLEANUP_DAYS > 0 else None,
        )
        db.add(file_source)
        await db.flush()  # Get the ID

        # Update table name with actual ID
        file_source.duckdb_table_name = self._generate_table_name(
            file.filename, file_source.id
        )

        file_path = None
        try:
            # Save file to disk
            file_path, file_hash, file_size = await self.save_file(
                file, session_id, is_global
            )

            # Update file source with path and size
            file_source.file_path = str(file_path)
            file_source.file_hash = file_hash
            file_source.file_size_bytes = file_size

            # Infer schema using DuckDB
            schema = await self.infer_schema(file_path, file_type, sheet_name)
            file_source.schema_cache = schema
            file_source.schema_updated_at = datetime.now(timezone.utc)
            file_source.row_count = schema.get('row_count', 0)

            # Mark as ready
            file_source.processing_status = 'ready'

            await db.commit()
            await db.refresh(file_source)

            logger.info(
                f"Processed file upload: {file.filename} -> {file_source.duckdb_table_name} "
                f"({file_source.row_count} rows, {file_size} bytes)"
            )

            return file_source

        except Exception as e:
            logger.error(f"Error processing file upload: {e}")
            file_source.processing_status = 'error'
            file_source.processing_error = str(e)
            await db.commit()

            # Clean up partial file on failure
            if file_path and file_path.exists():
                try:
                    file_path.unlink()
                    logger.debug(f"Cleaned up partial file: {file_path}")
                except Exception as cleanup_error:
                    logger.warning(f"Failed to clean up partial file {file_path}: {cleanup_error}")

            raise

    async def validate_file(self, file: UploadFile) -> Tuple[bool, str]:
        """
        Validate an uploaded file.

        Args:
            file: The uploaded file to validate

        Returns:
            Tuple of (is_valid, error_message)
        """
        # Check filename
        if not file.filename:
            return False, "Filename is required"

        # Check extension
        ext = Path(file.filename).suffix.lower()
        if ext not in self.allowed_extensions:
            return False, f"File type '{ext}' is not allowed. Allowed types: {', '.join(self.allowed_extensions)}"

        # Read a small header for content-type validation, then measure
        # total size by chunked reading to avoid loading the entire file
        # (up to 100MB) into memory just for validation.
        await file.seek(0)
        header = await file.read(4096)

        if len(header) == 0:
            await file.seek(0)
            return False, "File is empty"

        # Validate content type using magic bytes (basic check)
        try:
            is_valid_content = await self._validate_content(header, ext)
            if not is_valid_content:
                await file.seek(0)
                return False, f"File content does not match expected type for {ext}"
        except Exception as e:
            logger.warning(f"Content validation failed: {e}")
            # Continue if content validation fails - DuckDB will catch issues

        # Check file size by reading in chunks (avoids full file in memory)
        file_size = len(header)
        while True:
            chunk = await file.read(1024 * 1024)  # 1MB chunks
            if not chunk:
                break
            file_size += len(chunk)
            if file_size > self.max_size_bytes:
                await file.seek(0)
                return False, f"File size ({file_size / 1024 / 1024:.1f}MB) exceeds limit ({self.settings.FILE_MAX_SIZE_MB}MB)"

        await file.seek(0)  # Reset for later use

        return True, ""

    async def _validate_content(self, content: bytes, ext: str) -> bool:
        """Validate file content matches expected type."""
        # Check magic bytes for Excel files
        if ext in ('.xlsx', '.xls'):
            # XLSX files start with PK (ZIP format)
            if ext == '.xlsx':
                return content[:2] == b'PK'
            # XLS files have specific magic bytes
            if ext == '.xls':
                return content[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'

        # CSV files are text with delimiter-separated fields
        if ext == '.csv':
            try:
                sample = content[:4096].decode('utf-8')
                # Must have at least one line with a comma or tab delimiter
                lines = sample.splitlines()
                if not lines:
                    return False
                # Check that the first non-empty line contains a delimiter
                for line in lines:
                    stripped = line.strip()
                    if stripped:
                        has_comma = ',' in stripped
                        has_tab = '\t' in stripped
                        if has_comma or has_tab:
                            return True
                        # Single-column CSV is still valid but rare;
                        # accept if it looks like plain text (not binary)
                        return stripped.isprintable()
                return False
            except UnicodeDecodeError:
                return False

        return True

    async def save_file(
        self,
        file: UploadFile,
        session_id: Optional[str],
        is_global: bool,
    ) -> Tuple[Path, str, int]:
        """
        Save uploaded file to disk.

        Args:
            file: The uploaded file
            session_id: Chat session ID
            is_global: Whether file is global

        Returns:
            Tuple of (file_path, file_hash, file_size)
        """
        # Sanitize filename
        safe_filename = self._sanitize_filename(file.filename)

        # Determine storage path
        if is_global:
            storage_dir = self.upload_dir / 'global'
        elif session_id:
            storage_dir = self.upload_dir / 'sessions' / session_id
            storage_dir.mkdir(parents=True, exist_ok=True)
        else:
            storage_dir = self.upload_dir / 'global'

        # Stream file to disk while computing hash to avoid loading
        # the entire file (up to 100MB) into memory at once.
        await file.seek(0)
        hasher = hashlib.sha256()
        file_size = 0

        # Write to a temp file first, then rename once hash is known
        fd, tmp_path = tempfile.mkstemp(dir=str(storage_dir))
        try:
            async with aiofiles.open(fd, 'wb', closefd=True) as f:
                while True:
                    chunk = await file.read(1024 * 1024)  # 1MB chunks
                    if not chunk:
                        break
                    hasher.update(chunk)
                    await f.write(chunk)
                    file_size += len(chunk)
        except Exception:
            os.unlink(tmp_path)
            raise

        file_hash = hasher.hexdigest()

        # Use hash in filename to handle duplicates
        stored_filename = f"{file_hash[:12]}_{safe_filename}"
        file_path = storage_dir / stored_filename

        # Rename temp file to final path
        os.replace(tmp_path, str(file_path))

        logger.debug(f"Saved file to {file_path} ({file_size} bytes)")

        return file_path, file_hash, file_size

    async def infer_schema(
        self,
        file_path: Path,
        file_type: str,
        sheet_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Infer schema from file using DuckDB.

        Args:
            file_path: Path to the file
            file_type: Type of file (csv, xlsx, xls)
            sheet_name: Excel sheet name (optional)

        Returns:
            Dict with columns, row_count, and sample_values
        """
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(
            None,
            self._infer_schema_sync,
            str(file_path),
            file_type,
            sheet_name,
        )

    @contextlib.contextmanager
    def _duckdb_read_context(self, file_path: str, file_type: str, sheet_name: Optional[str]):
        """Context manager for DuckDB file reads with connection/cleanup management.

        Handles: DuckDB connection lifecycle, path validation, Excel-to-CSV
        conversion, and temp file cleanup.

        Yields:
            Tuple of (conn, read_query) - the DuckDB connection and the
            read_csv_auto query string for the file.
        """
        conn = duckdb.connect(':memory:')
        validated_path = validate_file_path(file_path, self.upload_dir)
        temp_csv_path = None
        try:
            if file_type in ('xlsx', 'xls'):
                safe_sheet = sanitize_sheet_name(sheet_name)
                temp_csv_path = excel_to_temp_csv(validated_path, safe_sheet)
                safe_csv_path = temp_csv_path.replace("'", "''")
                read_query = f"SELECT * FROM read_csv_auto('{safe_csv_path}', header=true)"
            else:
                safe_path = validated_path.replace("'", "''")
                read_query = f"SELECT * FROM read_csv_auto('{safe_path}', header=true)"
            yield conn, read_query
        finally:
            conn.close()
            if temp_csv_path:
                try:
                    os.unlink(temp_csv_path)
                except OSError:
                    pass

    def _infer_schema_sync(
        self,
        file_path: str,
        file_type: str,
        sheet_name: Optional[str],
    ) -> Dict[str, Any]:
        """Synchronous schema inference using DuckDB.

        Security: Validates file path is within upload directory and sanitizes
        sheet name to prevent SQL injection.
        """
        with self._duckdb_read_context(file_path, file_type, sheet_name) as (conn, read_query):
            # Get column info
            result = conn.execute(f"{read_query} LIMIT 0")
            columns = []
            for desc in result.description:
                col_type = str(desc[1]).upper()
                # Normalize type names
                if 'VARCHAR' in col_type or 'STRING' in col_type:
                    col_type = 'VARCHAR'
                elif 'INT' in col_type:
                    col_type = 'INTEGER'
                elif 'DOUBLE' in col_type or 'FLOAT' in col_type or 'REAL' in col_type:
                    col_type = 'DOUBLE'
                elif 'BOOL' in col_type:
                    col_type = 'BOOLEAN'
                elif 'DATE' in col_type:
                    col_type = 'DATE'
                elif 'TIME' in col_type and 'STAMP' in col_type:
                    col_type = 'TIMESTAMP'

                columns.append({
                    'name': desc[0],
                    'type': col_type,
                    'nullable': True,
                })

            # Get row count
            count_result = conn.execute(f"SELECT COUNT(*) FROM ({read_query})").fetchone()
            row_count = count_result[0] if count_result else 0

            # Get sample values for each column
            sample_values = {}
            sample_result = conn.execute(f"{read_query} LIMIT 5")
            rows = sample_result.fetchall()

            for i, col in enumerate(columns):
                values = []
                for row in rows:
                    val = row[i]
                    if val is not None:
                        # Convert to JSON-serializable format
                        if hasattr(val, 'isoformat'):
                            val = val.isoformat()
                        values.append(val)
                sample_values[col['name']] = values[:5]
                col['sample_values'] = values[:5]

            return {
                'columns': columns,
                'row_count': row_count,
                'sample_values': sample_values,
            }

    async def get_excel_sheets(self, file: UploadFile) -> List[str]:
        """
        Get list of sheets from an Excel file.

        Args:
            file: The uploaded Excel file

        Returns:
            List of sheet names
        """
        # Read file content
        await file.seek(0)
        content = await file.read()
        await file.seek(0)

        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(
            None,
            self._get_sheets_sync,
            content,
        )

    def _get_sheets_sync(self, content: bytes) -> List[str]:
        """Synchronously get sheet names using openpyxl."""
        import io
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        except Exception as e:
            logger.warning(f"Failed to read Excel sheets: {e}")
            return ['Sheet1']

    async def get_preview(
        self,
        file_source: FileSource,
        limit: int = 20,
    ) -> Dict[str, Any]:
        """
        Get preview rows from a file source.

        Args:
            file_source: The file source to preview
            limit: Maximum number of rows to return

        Returns:
            Dict with columns, data, row_count, total_rows, truncated
        """
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(
            None,
            self._get_preview_sync,
            file_source.file_path,
            file_source.file_type,
            file_source.sheet_name,
            limit,
        )

    def _get_preview_sync(
        self,
        file_path: str,
        file_type: str,
        sheet_name: Optional[str],
        limit: int,
    ) -> Dict[str, Any]:
        """Synchronously get preview data.

        Security: Validates file path is within upload directory and sanitizes
        sheet name to prevent SQL injection.
        """
        with self._duckdb_read_context(file_path, file_type, sheet_name) as (conn, read_query):
            # Get total count
            count_result = conn.execute(f"SELECT COUNT(*) FROM ({read_query})").fetchone()
            total_rows = count_result[0] if count_result else 0

            # Get preview rows
            result = conn.execute(f"{read_query} LIMIT {limit}")
            columns = [desc[0] for desc in result.description]
            rows = result.fetchall()

            data = []
            for row in rows:
                row_dict = {}
                for i, col in enumerate(columns):
                    val = row[i]
                    # Convert to JSON-serializable
                    if hasattr(val, 'isoformat'):
                        val = val.isoformat()
                    row_dict[col] = val
                data.append(row_dict)

            return {
                'columns': columns,
                'data': data,
                'row_count': len(data),
                'total_rows': total_rows,
                'truncated': len(data) < total_rows,
            }

    async def refresh_schema(
        self,
        file_source: FileSource,
        db: AsyncSession,
    ) -> FileSource:
        """
        Refresh schema for a file source.

        Args:
            file_source: The file source to refresh
            db: Database session

        Returns:
            Updated FileSource
        """
        try:
            schema = await self.infer_schema(
                Path(file_source.file_path),
                file_source.file_type,
                file_source.sheet_name,
            )

            file_source.schema_cache = schema
            file_source.schema_updated_at = datetime.now(timezone.utc)
            file_source.row_count = schema.get('row_count', 0)
            file_source.processing_status = 'ready'
            file_source.processing_error = None

            await db.commit()
            await db.refresh(file_source)

            logger.info(f"Refreshed schema for file source {file_source.id}")
            return file_source

        except Exception as e:
            logger.error(f"Error refreshing schema: {e}")
            file_source.processing_status = 'error'
            file_source.processing_error = str(e)
            await db.commit()
            raise

    async def cleanup_file(
        self,
        file_source: FileSource,
        db: AsyncSession,
    ) -> None:
        """
        Soft-delete a file source and remove its physical file.

        The database record is preserved with processing_status='deleted'
        so that chat sessions referencing this file can show it as removed
        instead of silently losing it.

        Args:
            file_source: The file source to delete
            db: Database session
        """
        # Delete physical file
        try:
            file_path = Path(file_source.file_path)
            if file_path.exists():
                file_path.unlink()
                logger.debug(f"Deleted file: {file_path}")
        except Exception as e:
            logger.warning(f"Failed to delete file {file_source.file_path}: {e}")

        # Soft-delete: mark as inactive/deleted instead of removing the record
        file_source.is_active = False
        file_source.processing_status = 'deleted'
        file_source.processing_error = 'File has been removed'
        await db.commit()

        logger.info(f"Cleaned up file source {file_source.id}: {file_source.name}")

    def _generate_table_name(self, filename: str, file_id: int) -> str:
        """
        Generate a unique DuckDB table name.

        Format: file_{id}_{sanitized_name}
        """
        # Sanitize name for SQL identifier
        safe_name = self._sanitize_filename(filename)
        safe_name = Path(safe_name).stem  # Remove extension
        safe_name = re.sub(r'[^a-z0-9_]', '_', safe_name.lower())
        safe_name = re.sub(r'_+', '_', safe_name)  # Collapse multiple underscores
        safe_name = safe_name[:30]  # Limit length

        return f"file_{file_id}_{safe_name}"

    def _sanitize_filename(self, filename: str) -> str:
        """
        Sanitize a filename to prevent path traversal and other issues.

        Args:
            filename: The original filename

        Returns:
            Sanitized filename
        """
        # Remove path components
        filename = os.path.basename(filename)

        # Remove null bytes and control characters
        filename = re.sub(r'[\x00-\x1f\x7f]', '', filename)

        # Remove path traversal attempts
        filename = filename.replace('..', '_')
        filename = filename.replace('/', '_')
        filename = filename.replace('\\', '_')

        # Replace problematic characters
        filename = re.sub(r'[<>:"|?*]', '_', filename)

        # Ensure filename is not empty
        if not filename or filename.startswith('.'):
            filename = 'unnamed_file'

        return filename


async def get_file_source_by_id(
    file_id: int,
    db: AsyncSession,
) -> Optional[FileSource]:
    """
    Get a file source by ID.

    Args:
        file_id: The file source ID
        db: Database session

    Returns:
        FileSource or None
    """
    result = await db.execute(
        select(FileSource).where(FileSource.id == file_id)
    )
    return result.scalar_one_or_none()


async def cleanup_expired_files(db: AsyncSession) -> int:
    """
    Delete file sources that have passed their expiration date.

    Removes both the physical files on disk and database records.

    Args:
        db: Database session

    Returns:
        Number of expired files cleaned up
    """
    from sqlalchemy import and_

    now = datetime.now(timezone.utc)
    stmt = select(FileSource).where(
        and_(
            FileSource.expires_at.isnot(None),
            FileSource.expires_at < now,
            FileSource.is_active.is_(True),
        )
    )

    result = await db.execute(stmt)
    expired_files = list(result.scalars().all())

    if not expired_files:
        return 0

    cleaned = 0
    handler = FileSourceHandler()

    # Local import to avoid circular dependency
    from src.core.file_source_session import FileSourceDuckDBSession

    for file_source in expired_files:
        try:
            # Unload from DuckDB before cleanup to free in-memory tables
            await FileSourceDuckDBSession.unload_table(file_source.duckdb_table_name)
            await handler.cleanup_file(file_source, db)
            cleaned += 1
            logger.info(f"Cleaned up expired file source {file_source.id}: {file_source.name}")
        except Exception as e:
            logger.error(f"Failed to clean up expired file {file_source.id}: {e}")

    if cleaned:
        logger.info(f"Expired file cleanup: removed {cleaned}/{len(expired_files)} files")

    return cleaned


async def list_file_sources(
    db: AsyncSession,
    session_id: Optional[str] = None,
    include_global: bool = True,
    user_id: Optional[str] = None,
    status: Optional[str] = None,
) -> List[FileSource]:
    """
    List file sources with optional filters.

    Args:
        db: Database session
        session_id: Filter by chat session ID
        include_global: Include global file sources
        user_id: Filter by user ID
        status: Filter by processing status

    Returns:
        List of FileSource records
    """
    from sqlalchemy import or_

    query = select(FileSource).where(FileSource.is_active.is_(True))

    # Build filters
    filters = []

    if session_id:
        if include_global:
            filters.append(or_(
                FileSource.chat_session_id == session_id,
                FileSource.is_global.is_(True),
            ))
        else:
            filters.append(FileSource.chat_session_id == session_id)
    elif include_global:
        filters.append(FileSource.is_global.is_(True))

    if user_id:
        filters.append(FileSource.user_id == user_id)

    if status:
        filters.append(FileSource.processing_status == status)

    if filters:
        query = query.where(*filters)

    query = query.order_by(FileSource.created_at.desc())

    result = await db.execute(query)
    return list(result.scalars().all())
